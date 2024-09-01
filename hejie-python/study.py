import sys
import webbrowser
import logging
import pyperclip
from http import client as httpclient
import json
from random import *
import re
import os
import pyinputplus as pyip
from pathlib import Path
import shelve
import pprint
import test
import shutil
import zipfile
import requests
from bs4 import BeautifulSoup
from urllib.parse import quote, urljoin, urlparse
import openpyxl
import subprocess
import time
import threading
import requests
import pyautogui
# from selenium import webdriver
# from selenium.webdriver.chrome.options import Options
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC

# # logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s-　%(message)s') 

a = 'start'
# wh =  pyautogui.size() 
# print(wh[0])

# for i in range(10):
#     pyautogui.move(100, 0, duration=0.25)  # right
#     pyautogui.move(0, 100, duration=0.25)  # down
#     pyautogui.move(-100, 0, duration=0.25) # left
#     pyautogui.move(0, -100, duration=0.25) # up
# pyautogui.mouseInfo()
b = pyautogui.getActiveWindow()

# def takeANap(stime):
#     print('sleep %s seconds...' % (stime))
#     time.sleep(stime)
#     print('Wake up!')

# # threadObj = threading.Thread(target=print, args=['Cats', 'Dogs', 'Frogs'],
# # kwargs={'sep': ' ———— ', 'end':'hejie'})
# # threadObj.start()

# downloadThreads = []               # a list of all the Thread objects
# for i in range(0, 14):      # loops 14 times, creates 14 threads
#     downloadThread = threading.Thread(target=takeANap, args=([i]))
#     downloadThreads.append(downloadThread)
#     downloadThread.start()
# # Wait for all threads to end.
# for downloadThread in downloadThreads:
#     downloadThread.join()
# print('全部线程完成')

# wb = openpyxl.load_workbook('demoexcel.xlsx')
# sheet = wb['test']
# for i in sheet['A:D']:
#     for j in i:
#         print(j.value)

# a = sheet.max_column
# a = sheet.max_row
# a = sheet.cell(row=1, column=2)
print(a)

# 调用浏览器进行自动化脚本
# 检查是否有 Chrome 实例在调试端口运行
# def is_chrome_running(port):
#     try:
#         response = requests.get(f"http://127.0.0.1:{port}/json")
#         return response.status_code == 200
#     except requests.ConnectionError:
#         return False

# # 调试端口号
# DEBUGGING_PORT = 9230

# # 启动 Chrome 并启用远程调试端口（如果没有已运行的实例）
# chrome_process = None
# if not is_chrome_running(DEBUGGING_PORT):
#     chrome_path = "/Applications/Google Chrome Dev.app/Contents/MacOS/Google Chrome Dev"
#     user_data_dir = "/Users/game-netease/Library/Application Support/Google/Chrome"

    # chrome_command = [
    #     chrome_path,
    #     f"--remote-debugging-port={DEBUGGING_PORT}",
    #     f"--user-data-dir={user_data_dir}",
    #     "--disable-gpu",
    #     "--disable-software-rasterizer"
    # ]

    # # 启动 Chrome
    # chrome_process = subprocess.Popen(chrome_command)

#     # 等待一段时间以确保 Chrome 启动完成
#     time.sleep(2)

# try:
#     print('开始执行')

#     # 设置Chrome浏览器选项
#     chrome_options = Options()
#     chrome_options.add_experimental_option("debuggerAddress", f"127.0.0.1:{DEBUGGING_PORT}")  # 连接到远程调试端口

#     # 指定ChromeDriver路径
#     chrome_driver_path = "/usr/local/bin/chromedriver/chromedriver"  # 确保路径正确
#     driver = webdriver.Chrome(executable_path=chrome_driver_path, options=chrome_options)

#     # 打开本地页面
#     local_url = "http://127.0.0.1:3000/homepage"
#     driver.get(local_url)

#     # 等待一段时间以观察是否重定向
#     time.sleep(2)

#     print(driver.current_url, local_url)

#     # 检查当前URL是否仍然是本地URL
#     if driver.current_url != local_url:
#         print("登录信息过期了，开始执行覆写操作")
        
#         # 打开指定的URL以获取cookies
#         url = "https://dreammaker-test.netease.com/homepage"
#         driver.get(url)

#         # 等待页面加载
#         time.sleep(2)  # 根据实际情况调整等待时间

#         # 获取当前的所有 cookies
#         cookies = driver.get_cookies()
#         # 修改特定的 cookies 的 domain 为 127.0.0.1 并设置
#         for cookie in cookies:
#             if cookie['name'] in ['ACCESS_TOKEN', 'AUTH_TOKEN', 'AUTH_USER']:
#                 new_cookie = cookie.copy()
#                 print(new_cookie)
#                 new_cookie['domain'] = '127.0.0.1'
#                 # 使用 JavaScript 创建和设置 cookie
#                 # driver.execute_script(f"document.cookie = '{new_cookie['name']}={new_cookie['value']}; domain={new_cookie['domain']}; path=/';")
#                 driver.add_cookie(new_cookie)
#         driver.get(local_url)

#         # 验证 cookie 是否设置成功
#         # for cookie in driver.get_cookies():
#         #     print(cookie)
#     else:
#         print("登录信息未过期，无需覆写")

#     # 关闭并退出浏览器
#     driver.quit()

# except Exception as e:
#     print(f"An error occurred: {e}")

# finally:
#     # 如果新启动了 Chrome 进程，关闭它
#     if chrome_process is not None:
#         chrome_process.terminate()


# 抓取图片
# 指定要抓取的URL
# url = "https://www.58pic.com/tupian/ceshiyongtu.html"

# # 发送HTTP请求并获取响应
# response = requests.get(url)
# response.raise_for_status()  # 检查请求是否成功

# # 解析HTML内容
# soup = BeautifulSoup(response.text, 'html.parser')

# # 创建保存图片的文件夹
# output_folder = Path.cwd() / "pachong"
# output_folder.mkdir(exist_ok=True)

# # 找到所有的img标签
# img_tags = soup.select('img')

# try:
#     for img in img_tags:
#         # 获取img标签中的src属性
#         img_url = img.get('src')
#         if img_url:
#             # 处理相对URL
#             img_url = urljoin(url, img_url)
            
#             # 解析URL以获取图片名称
#             img_name = Path(urlparse(img_url).path).name
            
#             # 下载图片并保存
#             img_response = requests.get(img_url)
#             img_response.raise_for_status()  # 检查请求是否成功
            
#             img_path = output_folder / img_name
#             with open(img_path, 'wb') as img_file:
#                 img_file.write(img_response.content)
#             print(f"Downloaded {img_name}")
#     print("All images have been downloaded.")
# except Exception as e:
#     print(e)

# try:
#     res = requests.get('https://www.epubit.com/onlineEbookReader?id=3c3aed65-eea0-4ba1-b0e2-4443333c1dc2&pid=df8d2065-0652-4c04-9043-1b8a80c49e68&isFalls=true&src=normal')
#     res.raise_for_status()
#     file = open('testhejie.txt', 'wb')
#     for chunk in res.iter_content(100000):
#         file.write(chunk)
#     file.close()
# except Exception as e:
#     print(e)


# 　&lt;class 'requests.models.Response'>
# ❷ >>> res.status_code == requests.codes.ok


# webbrowser.open('http://inventwithpython.com/')

# logging.debug('start')
# logging.disable()

# # for filename in Path().resolve().glob('*.py'):
# #     # os.unlink(filename)
# #     print(filename)

# a = Path.cwd()
# a = os.path.abspath(os.path.join(a, '../../'))
# # 约等于
# print(Path.cwd())
# shutil.copytree(a / 'lib', a /'lib2')
# shutil.move('test.txt', Path.cwd() / 'hejie/zone/aaa.txt')



# a=len(range(10))
# a=int(100)*7
# a =  2 + 2 != 5
# a = randint(0,10)

# # def spam():
# #     print(eggs)
# # eggs = 42
# # spam()


# spam={'a':12,'b':34,'c':56,'d':78}
# a = tuple(spam.items())

# b = [1,2,3,4,5,6,7]
# a = [k for k in b if k % 2 == 0]
# a = tuple(a)
# print(a)


# theBoard = {'top-L': 'O', 'top-M': 'O', 'top-R': 'O', 'mid-L': 'X', 'mid-M':
# 'X', 'mid-R': ' ', 'low-L': ' ', 'low-M': ' ', 'low-R': 'X'}

# phoneNumRegex = re.compile(r'\d\d\d-\d\d\d-\d\d\d\d')
# mo = phoneNumRegex.search('My number is 415-555-4242.')
# print('Phone number found: ' + mo.group())

#! python3
# pw.py - An insecure password locker program.

# 剪贴板的案例
# import sys, pyperclip

# PASSWORDS = {'email': 'F7minlBDDuvMJuxESSKHFhTxFtjVB6',
#              'blog': 'VmALvQyKAxiVH5G8v01if1MLZF3sdt',
#              'luggage': '12345'}

# if len(sys.argv) < 2:
#     print('Usage: py pw.py [account] - copy account password')
#     sys.exit()

# account = sys.argv[1] # first command line arg is the account name

# if account in PASSWORDS:
#     pyperclip.copy(PASSWORDS[account])
#     print('Password for ' + account + ' copied to clipboard.')
# else:
#     print('There is no account named ' + account)

# ! python3
# bulletPointAdder.py - Adds Wikipedia bullet points to the start
# of each line of text on the clipboard.

# text = pyperclip.paste()

# # Separate lines and add stars.
# lines = text.split('\n')
# for i in range(len(lines)):    # loop through all indexes for "lines" list
#     lines[i] = '* ' + lines[i]  # add star to each string in "lines" list
# text = '\n'.join(lines)
# pyperclip.copy(text)
# res = pyip.inputInt('请输入:', min=4, lessThan=10)


# def addsUpToTen(numbers):
#     numbersList = list(numbers)
#     for i, digit in enumerate(numbersList):
#         numbersList[i] = int(digit)
#     if sum(numbersList) != 10:
#         raise Exception('The digits must add up to 10, not %s.' %
#                         (sum(numbersList)))
#     return int(numbers)  # Return an int form of numbers.


# # res = pyip.inputCustom(addsUpToTen, '请输入:', limit=2, default=0)
# res = ''
# while res != 'no':
#     res = pyip.inputYesNo('请输入yes Or no:', yesVal='yes', noVal='no')
#     print(res)

# a = os.path.abspath('./lib/创作平台所有的翻译v2.xlsx')

# print(os.path.getsize(a), a)
# print(Path.cwd())
# a = open('./test.txt', 'a')
# a.write('nihao')
# a.close()
# print(open('./test.txt').read())
# shelfFile = shelve.open('mydata')
# cats = [{'name': 'hejie'}, {'name': 'fenfen'}]
# pcats = pprint.pformat(cats)
# file = open('test.py', 'w')
# file.write('cats=' + pcats + '\n')
# file.close()
# print(test.cats)

# shelfFile['cats'] = cats
# # type(shelfFile)
# print(shelfFile['cats'])
# shelfFile.close()

# shutil.move('./hejie12/sdf', './hejie')
# for i in list(Path().glob('*')):
#     print(i)

# a = Path.joinpath(Path.cwd(), 'hejie')

# a = zipfile.ZipFile('compressor.zip')
# print(a.namelist())
# a.extractall('hejie')

#!/usr/bin/env python
# -*- coding: UTF-8 -*-

# print(a)

# market_2nd = {'ns': 'green', 'ew': 'red'}
# mission_16th = {'ns': 'red', 'ew': 'green'}
# print(market_2nd.keys())


# # def switchLights(stoplight):
# #     for key in stoplight.keys():
# #         if stoplight[key] == 'green':
# #             stoplight[key] = 'yellow'
# #         if stoplight[key] == 'yellow':
# #             stoplight[key] = 'red'
# #         if stoplight[key] == 'red':
# #             stoplight[key] = 'green'
# #     assert 'red' in stoplight.values(), 'Neither light is red! ' + str(stoplight)


# # switchLights(market_2nd)
# # print(market_2nd)


# FORMAT = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
# logging.basicConfig(level=logging.DEBUG, format=FORMAT)

# logging.debug('Start of program')


# def factorial(n):
#     logging.debug('Start of factorial({})'.format(n))
#     total = 1
#     for i in range(n + 1):
#         total *= i
#         logging.debug('i is ' + str(i) + ', total is ' + str(total))
#     logging.debug('End of factorial({})'.format(n))
#     return total


# print(factorial(5))
# logging.debug('End of program')
# import logging

# 项目的完整代码如下
# print(sys.argv)

# if len(sys.argv) > 1:
#     # Get address from command line.
#     address = ' '.join(sys.argv[1:])
# else:
#     # Get address from clipboard.
#     address = pyperclip.paste()

# address = quote(address)

# webbrowser.open('https://ditu.amap.com/search?query=' + address)






